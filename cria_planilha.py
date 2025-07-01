import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

def criar_planilha_controle_impressoes():
    # Cria um novo workbook
    wb = openpyxl.Workbook()
    
    # Remove a planilha padrão e cria as abas necessárias
    del wb['Sheet']
    ws_impressoes = wb.create_sheet("Impressões")
    ws_resumo = wb.create_sheet("Resumo Professor Mensal")
    ws_cadastro = wb.create_sheet("Cadastro")

    ### 🔵 CONFIGURAÇÃO DE ESTILOS ###
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul
    header_font = Font(bold=True, color="FFFFFF")  # Texto branco
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal='center')
    total_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Laranja
    total_font = Font(bold=True)

    ### 📝 ABA "IMPRESSÕES" ###
    cabecalho_impressoes = ["Data", "Nome", "P&B", "Colorido", "Observações", "Total"]
    ws_impressoes.append(cabecalho_impressoes)

    # Formatação do cabeçalho
    for col in range(1, len(cabecalho_impressoes) + 1):
        celula = ws_impressoes.cell(row=1, column=col)
        celula.fill = header_fill
        celula.font = header_font
        celula.border = border
        celula.alignment = center_aligned
        ws_impressoes.column_dimensions[get_column_letter(col)].width = 18

    # Formatação de data e fórmulas
    for row in range(2, 1000):
        # Formata a coluna de data como DD/MM/AAAA
        ws_impressoes.cell(row=row, column=1).number_format = "DD/MM/YYYY"
        # Fórmula do total
        ws_impressoes.cell(row=row, column=6, value=f"=C{row}+D{row}")

    ### 📊 ABA "RESUMO PROFESSOR MENSAL" ###
    meses = ["Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    cabecalho_resumo = ["Professor"]
    for mes in meses:
        cabecalho_resumo.extend([f"P&B {mes}", f"Colorido {mes}"])
    ws_resumo.append(cabecalho_resumo)

    # Formatação do cabeçalho
    for col in range(1, len(cabecalho_resumo) + 1):
        celula = ws_resumo.cell(row=1, column=col)
        celula.fill = header_fill
        celula.font = header_font
        celula.border = border
        celula.alignment = center_aligned
        ws_resumo.column_dimensions[get_column_letter(col)].width = 14

    # Cores alternadas para linhas
    cor_linha_par = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    cor_linha_impar = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Adiciona professores e fórmulas
    for linha_prof in range(2, 100):
        nome_prof = f"=Cadastro!A{linha_prof - 1}"
        ws_resumo.cell(row=linha_prof, column=1, value=nome_prof)

        # Aplica cor alternada
        fill = cor_linha_par if linha_prof % 2 == 0 else cor_linha_impar
        for col in range(1, len(cabecalho_resumo) + 1):
            ws_resumo.cell(row=linha_prof, column=col).fill = fill
            ws_resumo.cell(row=linha_prof, column=col).border = border

        # Fórmulas para cada mês
        for idx_mes, mes in enumerate(meses):
            mes_num = idx_mes + 6
            
            # Fórmula P&B
            formula_pb = f'=SUMIFS(Impressões!C:C, Impressões!B:B, A{linha_prof}, Impressões!A:A, ">="&DATE(2025,{mes_num},1), Impressões!A:A, "<="&EOMONTH(DATE(2025,{mes_num},1),0))'
            ws_resumo.cell(row=linha_prof, column=2 + idx_mes * 2, value=formula_pb)
            
            # Fórmula Colorido
            formula_color = f'=SUMIFS(Impressões!D:D, Impressões!B:B, A{linha_prof}, Impressões!A:A, ">="&DATE(2025,{mes_num},1), Impressões!A:A, "<="&EOMONTH(DATE(2025,{mes_num},1),0))'
            ws_resumo.cell(row=linha_prof, column=3 + idx_mes * 2, value=formula_color)

    # ADICIONANDO OS TOTAIS MENSAL (P&B e Colorido)
    linha_total = 102  # Linha para os totais
    ws_resumo.cell(row=linha_total, column=1, value="TOTAL GERAL").font = total_font
    
    for idx_mes, mes in enumerate(meses):
        # Total P&B
        col_pb = 2 + idx_mes * 2
        formula_total_pb = f'=SUM(B2:B101)'.replace("B", get_column_letter(col_pb))
        ws_resumo.cell(row=linha_total, column=col_pb, value=formula_total_pb)
        
        # Total Colorido
        col_color = 3 + idx_mes * 2
        formula_total_color = f'=SUM(C2:C101)'.replace("C", get_column_letter(col_color))
        ws_resumo.cell(row=linha_total, column=col_color, value=formula_total_color)
        
        # Formatação
        for col in [col_pb, col_color]:
            ws_resumo.cell(row=linha_total, column=col).fill = total_fill
            ws_resumo.cell(row=linha_total, column=col).font = total_font
            ws_resumo.cell(row=linha_total, column=col).border = border

    ### 📋 ABA "CADASTRO" ###
    ws_cadastro.append(["Professores e Funcionários"])
    ws_cadastro.column_dimensions['A'].width = 25
    ws_cadastro['A1'].fill = header_fill
    ws_cadastro['A1'].font = header_font
    ws_cadastro['A1'].border = border
    ws_cadastro['A1'].alignment = center_aligned

    # Adiciona exemplos
    professores_exemplo = ["Leiliane", "Carlos", "Ana", "Pedro", "Mariana"]
    for linha, nome in enumerate(professores_exemplo, start=2):
        ws_cadastro.cell(row=linha, column=1, value=nome)
        ws_cadastro.cell(row=linha, column=1).border = border

    ### 💾 SALVA O ARQUIVO ###
    nome_arquivo = f"Controle_Impressoes_Completo_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    wb.save(nome_arquivo)
    print(f"✅ Planilha criada com sucesso: {nome_arquivo}")
    print("👉 Adicione os dados na aba 'Impressões' e o resumo será atualizado automaticamente!")

if __name__ == "__main__":
    criar_planilha_controle_impressoes()